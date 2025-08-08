from flask import Flask, request
from flask_cors import CORS
import os
import openpyxl
from datetime import datetime

app = Flask(__name__)
CORS(app)

excel_file = excel_file = os.path.join(os.path.dirname(__file__), "game_data.xlsx")


# Create file with headers if not present
if not os.path.exists(excel_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "PlayerID", "Name", "Age", "Gender", "TimeTaken (s)",
        "PairsMatchedOnTime", "Mistakes", "WonOrLost",
        "Timestamp", "Speed"
    ])
    wb.save(excel_file)

@app.route("/save", methods=["POST"])
def save():
    data = request.json
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    player_id = ws.max_row  # auto-increment ID
    name = data.get("name")
    age = data.get("age")
    gender = data.get("gender")
    time_taken = data.get("time")
    matched = data.get("matches")
    mistakes = data.get("mismatches")
    status = data.get("status")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if time_taken > 40:
        speed = "Slow"
    elif time_taken >= 20:
        speed = "Moderate"
    else:
        speed = "Fast"

    ws.append([
        player_id, name, age, gender, time_taken,
        matched, mistakes, status, timestamp, speed
    ])
    wb.save(excel_file)

    return {"status": "success"}

if __name__ == "__main__":
    app.run(port=5000)
